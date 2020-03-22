import openpyxl, os
def spaces(word,spaces):
    return " "*(spaces-len(word)) if spaces >= len(word) else " "*(len(word)-spaces)
# logging.disable("WARN")
os.chdir("/home/ylli/Documents")
workbook = openpyxl.Workbook()
sheet = workbook.get_sheet_by_name("Sheet")
sheet.title = "Sheet"
for i in range(1,10):
    sheet.cell(column=1,row=i, value=i)
print(workbook.get_sheet_names())
print(sheet["A1"].value)
workbook.save("test.xlsx")
# workbook = openpyxl.load_workbook("Links to Tasks (2).xlsx")
# # print(type(workbook))
# # sheets = workbook.get_sheet_names()
# # print("Choose a sheet: ",*sheets)
# # enter = input("")
# sheet1 = workbook.get_sheet_by_name("Sheet1")
# # print(sheet1["A"])
# for names,sites in zip(sheet1["A"],sheet1["B"]):
#     print(f"{names.value}{spaces(names.value,15)} : {sites.value}")