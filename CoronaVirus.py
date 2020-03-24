#! /usr/bin/python3
import bs4,requests, openpyxl,os,re
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
os.chdir("/home/ylli/Documents")
site = requests.get("https://en.m.wikipedia.org/wiki/2019%E2%80%9320_coronavirus_pandemic_by_country_and_territory")
soup = bs4.BeautifulSoup(site.text,"html.parser")
countries = soup.select("#covid19-container > table > tbody > tr > th")
cases = soup.select("#covid19-container > table > tbody > tr > td")
excel = openpyxl.Workbook()
# old = openpyxl.load_workbook("CoronaVirus.x")
sheet = excel.active
sheet.merge_cells("A1:A2")
sheet["A1"],sheet["A1"].font = "Corona Virus",Font(color=colors.RED)
sheet["B2"]= (countries[7].text).strip()
sheet["B1"] = "Cases"
sheet["C2"] = (countries[8].text).strip()
sheet["C1"] = "Deaths"
sheet["D2"] = (countries[9].text).strip()
sheet["D1"] = "Recovered"
number = 3
sheet.column_dimensions["A"].width =sheet.column_dimensions["B"].width =sheet.column_dimensions["C"].width =sheet.column_dimensions["D"].width = 25
for i,j in enumerate(countries[12:], start=1):
    j = (j.text).strip()
    if j == "":
        continue
    j = re.sub("\[.*]","",j)
    sheet.cell(column=1,row=number,value=j)
    number+=1
number =3
starter = 0
for i,j in enumerate(cases):
    j = (j.text).strip()
    if j.startswith("As of"):
        break
    if j == "" or j.startswith("["):
        continue
    # if i % 4 == 0:
    sheet.cell(column=2+starter,row=number,value=j)
    starter+=1
    if starter ==3:
        starter = 0
        number+=1
# sheet.column_dimensions.width = 20
for i in sheet["A1":"C120"]:
    for j in i:
        print(j.value)
excel.save("CoronaVirus.xlsx")
choice = input("Do you wanna look at the data? ")
if choice.upper().startswith("Y"):
    os.system("xdg-open CoronaVirus.xlsx")